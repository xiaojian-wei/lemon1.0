# -*- coding: utf-8 -*-
# @Author:ASUS
# @File :method.py
# @Software :PyCharm
# @Time :2021/8/15 22:00

"""
接口自动化测试步骤
1.excel测试用例准备ok，代码自动去读取测试数据
2.发送接口请求，得到响应结果
3.断言：执行结果 vs 预期结果 --通过/不通过
4.写入最终结果 到excel表格（不涉及输出测试报告）
"""
from openpyxl import load_workbook
import requests


# 封装读excel函数
def read_case(file_name, sheet_name):
    # 打开excel文件
    wb = load_workbook(file_name)
    # 打开对应的sheet表单
    sh = wb[sheet_name]
    # 获取总行数
    max_row = sh.max_row
    # 获取数据（表头对应的是第一行）
    case_list = []
    for i in range(2, max_row + 1):
        dict1 = dict(
            case_id=sh.cell(row=i, column=1).value,
            url=sh.cell(row=i, column=5).value,
            data=sh.cell(row=i, column=6).value,
            expect=sh.cell(row=i, column=7).value
        )
        # 每次循环，把生成dict追加到list列表中
        case_list.append(dict1)
    return case_list


# 封装写入函数
def write(file_name, sheet_name, row, column, final_result):
    # 打开excel文件
    wb = load_workbook(file_name)
    # 打开对应的sheet表单
    sh = wb[sheet_name]
    # 写入
    sh.cell(row=row, column=column).value = final_result
    # 保存
    wb.save(file_name)


# 发送接口请求（post）
def api_func(url, data):
    headers = {"X-Lemonban-Media-Type": "lemonban.v2", "Content-Type": "application/json"}
    res = requests.post(url=url, json=data, headers=headers).json()
    return res